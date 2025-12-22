import time
import json
import re
import requests
from decimal import Decimal, InvalidOperation
from datetime import timedelta, datetime
from collections import defaultdict
from django.db import transaction
import secrets

import openpyxl

from django.conf import settings
from django.views.decorators.csrf import csrf_exempt
from django.shortcuts import get_object_or_404, render
from django.db.models import Count, Sum, Avg, Q, F, Value, DecimalField, Min
from django.db.models.functions import TruncHour, TruncDay, TruncWeek, TruncMonth, Coalesce
from django.utils import timezone

from rest_framework.decorators import (
    api_view, permission_classes, parser_classes, renderer_classes
)
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from rest_framework.renderers import JSONRenderer
from rest_framework.response import Response
from rest_framework import status

from .models import (
    CustomUser, Product, ProductVariant, CartItem, WishlistItem, Order, CATEGORIES, RecentlyViewed,
    OrderItem, Notification, Payment, Quotation,
    ProductConversation, ProductMessage, QuotationRequest, 
    Review, Invoice, PendingUser,BuyerProfile, SellerProfile, PasswordResetToken, ShippingAddress, BillingAddress
)
from .serializers import (
    RegisterSerializer, ProductSerializer, CartItemSerializer,
    WishlistItemSerializer, OrderSerializer, NotificationSerializer,
    QuotationSerializer, ProductConversationSerializer, ProductMessageSerializer,
    QuotationRequestSerializer, OrderStatusUpdateSerializer, ReviewSerializer,
    InvoiceSerializer, VerifyOTPSerializer, RecentlyViewedSerializer
)

from rest_framework.pagination import PageNumberPagination


from .utils.otp_utils import generate_otp, send_otp_email, send_password_reset_email, send_delivery_otp_email, send_order_confirmation_email, send_order_status_update_email, send_seller_new_order_email
from django.contrib.auth import get_user_model
import openpyxl

User = get_user_model()

# ==========================================================
# USER ENDPOINTS
# ==========================================================
def index(request):
    return render(request, "index.html")

@api_view(["POST"])
@permission_classes([AllowAny])
def register_buyer(request):
    """
    Step 1: Buyer submits basic details.
    Save to PendingUser + send OTP.
    """
    data = request.data
    

    serializer = RegisterSerializer(data=data)
    serializer.is_valid(raise_exception=True)

    name = data["name"]
    phone = data["phone"]
    email = data["email"]
    password = data["password"]
    role = "buyer"

    # Generate OTP
    otp = generate_otp()

    # Create or update pending user
    PendingUser.objects.update_or_create(
        email=email,
        defaults={
            "name": name,
            "phone": phone,
            "role": role,
            "password": password,
            "otp": otp,
        },
    )

    # Send OTP
    send_otp_email(email, otp)

    return Response({"message": "OTP sent to your email"}, status=200)

@api_view(["POST"])
@permission_classes([AllowAny])
def register_seller(request):
    """
    Seller registration — NO OTP. 
    Creates CustomUser + SellerProfile (pending approval).
    """
    data = request.data

    required_fields = [
        "name", "phone", "email",
        "businessName", "entityType",
        "gstNumber", "panNumber",
        "addressLine1", "city", "state", "pincode",
    ]

    for field in required_fields:
        if not data.get(field):
            return Response({"error": f"{field} is required"}, status=400)
    
    temporary_password = secrets.token_urlsafe(10)  # random secure password
   
    user = CustomUser.objects.create_user(
        email=data["email"],
        name=data["name"],
        phone=data["phone"],
        role="seller",
        password=temporary_password,
    )

    # Create Seller Profile
    profile = SellerProfile.objects.create(
        user=user,
        business_name=data.get("businessName"),
        entity_type=data.get("entityType"),
        gst_number=data.get("gstNumber"),
        pan_number=data.get("panNumber"),

        seller_categories=data.get("sellerCategories", []),
        designation=data.get("designation"),
        website_url=data.get("websiteUrl"),
        linkedin_url=data.get("linkedinUrl"),

        temp_password=temporary_password,


        address_line1=data.get("addressLine1"),
        address_line2=data.get("addressLine2"),
        city=data.get("city"),
        state=data.get("state"),
        pincode=data.get("pincode"),

        is_approved=False,
    )

    # Handle optional file uploads
    if "businessRegDoc" in request.FILES:
        profile.business_document = request.FILES["businessRegDoc"]

    if "gstCertificate" in request.FILES:
        profile.gst_certificate = request.FILES["gstCertificate"]

    profile.save()

    return Response(
        {"message": "Seller registered. Awaiting admin approval."},
        status=201,
    )



@api_view(["POST"])
@permission_classes([AllowAny])
def verify_otp_register(request):
    serializer = VerifyOTPSerializer(data=request.data)
    serializer.is_valid(raise_exception=True)

    email = serializer.validated_data["email"]
    otp = serializer.validated_data["otp"]
    

    try:
        pending = PendingUser.objects.get(email=email)
    except PendingUser.DoesNotExist:
        return Response({"error": "Invalid email"}, status=400)

    if pending.otp != otp:
        return Response({"error": "Incorrect OTP"}, status=400)

    if not pending.is_valid():
        pending.delete()
        return Response({"error": "OTP expired"}, status=400)

    # Create final buyer user
 

    user = CustomUser.objects.create_user(
        email=pending.email,
        name=pending.name,
        phone=pending.phone,
        role="buyer",
        password=pending.password,
    )
    

    BuyerProfile.objects.create(
        user=user,
        username=request.data.get("username"),
        buyer_category=request.data.get("buyerCategory"),
        organization_name=request.data.get("organizationName"),
        city=request.data.get("city"),
        state=request.data.get("state"),
        pincode=request.data.get("pincode"),
    )

    pending.delete()
    return Response({"message": "Registration successful"}, status=201)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def user_profile(request):
    user = request.user

    base_data = {
        "name": user.name,
        "email": user.email,
        "phone": user.phone,
        "role": user.role,
    }


    if user.role == "buyer":
        try:
            bp = user.buyer_profile
            base_data.update({
                "username": bp.username,
                "buyerCategory": bp.buyer_category,
                "organizationName": bp.organization_name,
                "city": bp.city,
                "state": bp.state,
                "pincode": bp.pincode,
            })
        except BuyerProfile.DoesNotExist:
            pass


    elif user.role == "seller":
        try:
            sp = user.seller_profile
            base_data.update({
                "businessName": sp.business_name,
                "entityType": sp.entity_type,
                "gstNumber": sp.gst_number,
                "panNumber": sp.pan_number,
                "sellerCategories": sp.seller_categories,
                "designation": sp.designation,
                "websiteUrl": sp.website_url,
                "linkedinUrl": sp.linkedin_url,
                "addressLine1": sp.address_line1,
                "addressLine2": sp.address_line2,
                "city": sp.city,
                "state": sp.state,
                "pincode": sp.pincode,
            })
        except SellerProfile.DoesNotExist:
            pass

    return Response(base_data)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def user_profile(request):
    user = request.user

    base_data = {
        "name": user.name,
        "email": user.email,
        "phone": user.phone,
        "role": user.role,
    }

    
    if user.role == "buyer":
        try:
            bp = user.buyer_profile
            base_data.update({
                "username": bp.username,
                "buyerCategory": bp.buyer_category,
                "organizationName": bp.organization_name,
                "city": bp.city,
                "state": bp.state,
                "pincode": bp.pincode,
                "shippingAddresses": [
                    {
                        "id": s.id,
                        "name": s.name,
                        "phone": s.phone,
                        "addressLine1": s.address_line1,
                        "addressLine2": s.address_line2,
                        "city": s.city,
                        "state": s.state,
                        "pincode": s.pincode,
                        "isDefault": s.is_default,
                    }
                    for s in user.shipping_addresses.all()
                ],
                "billingAddress": (
                    {
                        "name": user.billing_address.name,
                        "phone": user.billing_address.phone,
                        "addressLine1": user.billing_address.address_line1,
                        "addressLine2": user.billing_address.address_line2,
                        "city": user.billing_address.city,
                        "state": user.billing_address.state,
                        "pincode": user.billing_address.pincode,
                    }
                    if hasattr(user, "billing_address") else None
                )
            })
        except BuyerProfile.DoesNotExist:
            pass

    elif user.role == "seller":
        try:
            sp = user.seller_profile
            base_data.update({
                "businessName": sp.business_name,
                "entityType": sp.entity_type,
                "gstNumber": sp.gst_number,
                "panNumber": sp.pan_number,
                "sellerCategories": sp.seller_categories,
                "designation": sp.designation,
                "websiteUrl": sp.website_url,
                "linkedinUrl": sp.linkedin_url,
                "addressLine1": sp.address_line1,
                "addressLine2": sp.address_line2,
                "city": sp.city,
                "state": sp.state,
                "pincode": sp.pincode,
            })
        except SellerProfile.DoesNotExist:
            pass

    return Response(base_data)


# =================================================================
#                 UPDATE PROFILE (BUYER + SELLER)
# =================================================================

@api_view(["PUT", "PATCH"])
@permission_classes([IsAuthenticated])
@parser_classes([JSONParser, MultiPartParser, FormParser])
def update_user_profile(request):
    import json

    user: CustomUser = request.user
    data = request.data

    user.name = data.get("name", user.name)
    user.phone = data.get("phone", user.phone)

    new_email = data.get("email")
    if new_email and new_email != user.email:
        if CustomUser.objects.filter(email=new_email).exclude(id=user.id).exists():
            return Response({"error": "Email already in use"}, status=400)
        user.email = new_email

    user.save()

    if user.role == "buyer":
        profile = BuyerProfile.objects.get(user=user)

        # --- Basic buyer fields ---
        profile.username = data.get("username", profile.username)
        profile.buyer_category = data.get("buyerCategory", profile.buyer_category)
        profile.organization_name = data.get("organizationName", profile.organization_name)
        profile.city = data.get("city", profile.city)
        profile.state = data.get("state", profile.state)
        profile.pincode = data.get("pincode", profile.pincode)
        profile.save()

        raw_shipping = data.get("shippingAddresses")
        if isinstance(raw_shipping, str):
            shipping_addresses = json.loads(raw_shipping)
        else:
            shipping_addresses = raw_shipping or []

        existing_ids = {s.id for s in user.shipping_addresses.all()}
        incoming_ids = {
            a.get("id") for a in shipping_addresses
            if a.get("id") and str(a.get("id")).isdigit()
        }

        to_delete = existing_ids - incoming_ids
        if to_delete:
            user.shipping_addresses.filter(id__in=to_delete).delete()

      
        for addr in shipping_addresses:

            if not addr or not any(addr.values()):
                continue

            addr_id = addr.get("id")
            id_exists = (
                addr_id
                and str(addr_id).isdigit()
                and user.shipping_addresses.filter(id=addr_id).exists()
            )

            if not id_exists:
                obj = user.shipping_addresses.create(
                    user=user,
                    name=addr["name"],
                    phone=addr["phone"],
                    address_line1=addr["addressLine1"],
                    address_line2=addr.get("addressLine2"),
                    city=addr["city"],
                    state=addr["state"],
                    pincode=addr["pincode"],
                    is_default=addr.get("isDefault", False),
                )
            else:

                obj = user.shipping_addresses.get(id=addr_id)
                obj.name = addr.get("name", obj.name)
                obj.phone = addr.get("phone", obj.phone)
                obj.address_line1 = addr.get("addressLine1", obj.address_line1)
                obj.address_line2 = addr.get("addressLine2", obj.address_line2)
                obj.city = addr.get("city", obj.city)
                obj.state = addr.get("state", obj.state)
                obj.pincode = addr.get("pincode", obj.pincode)
                obj.is_default = addr.get("isDefault", obj.is_default)
                obj.save()

            if addr.get("isDefault") is True:
                user.shipping_addresses.exclude(id=obj.id).update(is_default=False)

        raw_billing = data.get("billingAddress")
        if isinstance(raw_billing, str):
            billing_data = json.loads(raw_billing)
        else:
            billing_data = raw_billing or {}

        

        if billing_data and any(billing_data.values()):
            try:
                bill = user.billing_address
            except BillingAddress.DoesNotExist:
                bill = BillingAddress(user=user)

            bill.name = billing_data.get("name", bill.name)
            bill.phone = billing_data.get("phone", bill.phone)
            bill.address_line1 = billing_data.get("addressLine1", bill.address_line1)
            bill.address_line2 = billing_data.get("addressLine2", bill.address_line2)
            bill.city = billing_data.get("city", bill.city)
            bill.state = billing_data.get("state", bill.state)
            bill.pincode = billing_data.get("pincode", bill.pincode)
            bill.save()

        shipping_list = [
            {
                "id": s.id,
                "name": s.name,
                "phone": s.phone,
                "addressLine1": s.address_line1,
                "addressLine2": s.address_line2,
                "city": s.city,
                "state": s.state,
                "pincode": s.pincode,
                "isDefault": s.is_default,
            }
            for s in user.shipping_addresses.all()
        ]

        billing_info = None
        if hasattr(user, "billing_address"):
            b = user.billing_address
            billing_info = {
                "name": b.name,
                "phone": b.phone,
                "addressLine1": b.address_line1,
                "addressLine2": b.address_line2,
                "city": b.city,
                "state": b.state,
                "pincode": b.pincode,
            }

        return Response({
            "message": "Buyer profile updated successfully",
            "shippingAddresses": shipping_list,
            "billingAddress": billing_info,
        })

    return Response({"error": "Invalid user role"}, status=400)


@api_view(["POST"])
@permission_classes([AllowAny])
def request_password_reset(request):
 
    email = request.data.get("email")

    if not email:
        return Response({"error": "Email required"}, status=400)

    try:
        user = CustomUser.objects.get(email=email)
      
    except CustomUser.DoesNotExist:
        return Response({"error": "Email not found"}, status=404)

    code = generate_otp()


    PasswordResetToken.objects.create(
        user=user,
        token=code
    )

    send_password_reset_email(email, code)

    return Response({"message": "Password reset code sent"}, status=200)

@api_view(["POST"])
@permission_classes([AllowAny])
def verify_password_reset(request):
    email = request.data.get("email")
    code = request.data.get("code")
    new_password = request.data.get("new_password")

    try:
        user = CustomUser.objects.get(email=email)
    except CustomUser.DoesNotExist:
        return Response({"error": "Email not found"}, status=404)

    try:
        token_obj = PasswordResetToken.objects.filter(user=user).latest("created_at")
    except PasswordResetToken.DoesNotExist:
        return Response({"error": "No reset requested"}, status=400)

    if token_obj.token != code:
        return Response({"error": "Invalid code"}, status=400)

    if not token_obj.is_valid():
        return Response({"error": "Code expired"}, status=400)

    user.set_password(new_password)
    user.save()

    token_obj.delete()

    return Response({"message": "Password reset successful"}, status=200)


# ==========================================================
# PRODUCT ENDPOINTS
# ==========================================================

class ProductPagination(PageNumberPagination):
    page_size = 20
    page_size_query_param = "page_size"
    max_page_size = 100



@api_view(["GET"])
@permission_classes([AllowAny])
def list_products(request):
    qs = (
        Product.objects
        .select_related("seller")
        .prefetch_related("variants", "reviews")
        .order_by("-created_at")
    )

    category = request.GET.get("category")
    seller = request.GET.get("seller")

    if category:
        qs = qs.filter(category=category)

    if seller:
        qs = qs.filter(seller_id=seller)

    if request.user.is_authenticated and request.user.role == "seller":
        qs = qs.filter(seller=request.user)

    paginator = ProductPagination()
    page = paginator.paginate_queryset(qs, request)

    serializer = ProductSerializer(
        page,
        many=True,
        context={"request": request}
    )

    return paginator.get_paginated_response(serializer.data)


@api_view(["POST"])
@permission_classes([IsAuthenticated])
@parser_classes([MultiPartParser, FormParser, JSONParser])
def add_product(request):
    if request.user.role != "seller":
        return Response({"detail": "Only sellers can add products."}, status=403)
    data = request.data.copy()
    variants = []
    if "variants" in data:
        raw = data.get("variants")
        try:
            variants = json.loads(raw) if isinstance(raw, str) else raw
        except json.JSONDecodeError:
            return Response({"variants": ["Invalid JSON format."]}, status=400)
    serializer = ProductSerializer(data=data, context={"request": request})
    if serializer.is_valid():
        product = serializer.save(seller=request.user)
        for v in variants:
            ProductVariant.objects.update_or_create(
                product=product,
                variant_label=v.get("variant_label"),
                defaults={
                    "est_price": v.get("est_price"),
                    "price": v.get("price"),
                    "discount": v.get("discount")
                }
            )

        return Response(ProductSerializer(product, context={"request": request}).data, status=201)
    return Response(serializer.errors, status=400)


@api_view(["POST"])
@permission_classes([IsAuthenticated])
@parser_classes([MultiPartParser, FormParser])
def upload_products(request):
    """
    Bulk upload products + variants via Excel file.
    Variant-level pricing & discount.
    Continues on errors and returns row-wise report.
    """



    excel_file = request.FILES.get("file")
    if not excel_file:
        return Response({"detail": "No file uploaded."}, status=400)

    try:
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
    except Exception as e:
        return Response({"detail": f"Invalid Excel file: {str(e)}"}, status=400)

    headers = [str(cell.value).strip() for cell in ws[1]]

    required_fields = [
        "name",
        "ref_no",
        "category",
        "brand",
        "variant_label",
        "est_price",
        "gst",
    ]

    for field in required_fields:
        if field not in headers:
            return Response(
                {"detail": f"Missing required column: '{field}'"},
                status=400,
            )

    index_map = {h: i for i, h in enumerate(headers)}

    product_cache = {}
    created_products = set()
    row_errors = []

    def to_decimal(value):
        if value in [None, "", " ", "nan"]:
            return None
        try:
            return Decimal(str(value))
        except InvalidOperation:
            return None

    valid_categories = [c[0] for c in CATEGORIES]


    for excel_row_index, row in enumerate(
        ws.iter_rows(min_row=2, values_only=True), start=2
    ):
        if not row or not any(row):
            continue

        def get(col):
            try:
                val = row[index_map[col]]
                return None if val in ["", " ", None] else val
            except Exception:
                return None

        try:
            ref_no = str(get("ref_no")).strip()
            if not ref_no:
                raise ValueError("Missing ref_no")


            if ref_no not in product_cache:
                category = (
                    str(get("category"))
                    .lower()
                    .replace(" ", "_")
                    .strip()
                )

                if category not in valid_categories:
                    raise ValueError(
                        f"Invalid category '{category}'. "
                        f"Allowed: {valid_categories}"
                    )

                gst_value = to_decimal(get("gst"))
                if gst_value is None:
                    raise ValueError("Invalid GST value")

                product_data = {
                    "name": get("name"),
                    "ref_no": ref_no,
                    "description": get("description"),
                    "category": category,
                    "hsn": get("hsn"),
                    "brand": get("brand"),
                    "cas_no": get("cas_no"),
                    "gst": gst_value,
                }

                serializer = ProductSerializer(
                    data=product_data,
                    context={"request": request},
                )

                if not serializer.is_valid():
                    raise ValueError(
                        f"Product validation failed: {serializer.errors}"
                    )

                product = serializer.save(seller=request.user)
                product_cache[ref_no] = product
                created_products.add(product.id)

            product = product_cache[ref_no]


            variant_label = get("variant_label")
            if not variant_label:
                raise ValueError("Missing variant_label")

            est_price = to_decimal(get("est_price"))
            if est_price is None:
                raise ValueError("Invalid est_price")

            price = to_decimal(get("price"))
            discount = to_decimal(get("discount"))  # nullable

            if discount is not None and (discount < 0 or discount > 100):
                raise ValueError("Discount must be between 0 and 100")

            ProductVariant.objects.update_or_create(
                product=product,
                variant_label=str(variant_label),
                defaults={
                    "est_price": est_price,
                    "price": price,
                    "discount": discount,
                },
            )

        except Exception as e:
            row_errors.append(
                {
                    "row": excel_row_index,
                    "error": str(e),
                }
            )
            continue

    return Response(
        {
            "created": len(created_products),
            "errors": row_errors,
        },
        status=201,
    )

@api_view(["PUT"])
@permission_classes([IsAuthenticated])
@parser_classes([MultiPartParser, FormParser, JSONParser])
def update_product(request, pk):
    try:
        product = Product.objects.get(pk=pk, seller=request.user)
    except Product.DoesNotExist:
        return Response({"detail": "Not found"}, status=404)
    data = request.data.copy()
    variants = []
    if "variants" in data:
        raw = data.get("variants")
        try:
            variants = json.loads(raw) if isinstance(raw, str) else raw
        except json.JSONDecodeError:
            return Response({"variants": ["Invalid JSON format."]}, status=400)
    serializer = ProductSerializer(product, data=data, partial=True, context={"request": request})
    if serializer.is_valid():
        updated = serializer.save()
        updated.variants.all().delete()
        for v in variants:
            print(v)
            variant, _ = ProductVariant.objects.update_or_create(
                product=updated,
                variant_label=v.get("variant_label"),
                defaults={
                    "est_price": v.get("est_price"),
                    "price": v.get("price"),
                    "discount": v.get("discount"),
                }
            )


        return Response(ProductSerializer(updated, context={"request": request}).data)
    return Response(serializer.errors, status=400)

@api_view(["DELETE"])
@permission_classes([IsAuthenticated])
def delete_product(request, pk):
    try:
        p = Product.objects.get(pk=pk, seller=request.user)
    except Product.DoesNotExist:
        return Response({"detail": "Not found"}, status=404)
    p.delete()
    return Response({"detail": "Product deleted."}, status=204)

# ----------------------------------------------------------------
# Recently Viewed
# ----------------------------------------------------------------
@api_view(["POST"])
@permission_classes([IsAuthenticated])
def add_recent_view(request, product_id):

    user = request.user

    try:
        product = Product.objects.get(id=product_id)
    
    except Product.DoesNotExist:
        return Response({"error": "Product not found"}, status=404)

    # Create or update timestamp
    entry, created = RecentlyViewed.objects.update_or_create(
        user=user,
        product=product,
        defaults={"viewed_at": timezone.now()}
    )

    # Keep only last 20 items
    recent_ids = (
        RecentlyViewed.objects.filter(user=user)
        .order_by("-viewed_at")
        .values_list("id", flat=True)[:20]
    )

    RecentlyViewed.objects.filter(user=user).exclude(id__in=list(recent_ids)).delete()

    return Response({"message": "Added to recently viewed"})

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_recently_viewed(request):
    user = request.user
    items = RecentlyViewed.objects.filter(user=user).select_related("product")
    data = RecentlyViewedSerializer(items, many=True, context={"request": request}).data
    
    return Response(data)

# ==========================================================
# CART & WISHLIST 
# ==========================================================
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def view_cart(request):
    items = CartItem.objects.filter(user=request.user).select_related("product", "user")
    return Response(CartItemSerializer(items, many=True).data)


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def add_to_cart(request):
    product_id = request.data.get("product_id")
    variant_id = request.data.get("variant_id")
    quantity = int(request.data.get("quantity", 1))

    product = get_object_or_404(Product, pk=product_id)
    variant = None

    if variant_id:
        variant = get_object_or_404(ProductVariant, pk=variant_id, product=product)
        price = variant.price if variant.price else variant.est_price or 0
    else:
        price = product.variants.first().price if product.variants.exists() else 0

    item, created = CartItem.objects.get_or_create(
        user=request.user,
        product=product,
        variant=variant,
        defaults={"quantity": quantity}
    )

    if not created:
        item.quantity = quantity
    item.save()

    return Response(CartItemSerializer(item).data, status=201)


@api_view(["DELETE"])
@permission_classes([IsAuthenticated])
def remove_from_cart(request, pk):
    try:
        item = CartItem.objects.get(pk=pk, user=request.user)
    except CartItem.DoesNotExist:
        return Response({"detail": "Not found"}, status=404)
    if item.quantity > 1:
        item.quantity -= 1
        item.save()
        return Response({"detail": f"Quantity decreased to {item.quantity}"})
    item.delete()
    return Response({"detail": "Removed from cart"})

@api_view(["DELETE"])
@permission_classes([IsAuthenticated])
def clear_from_cart(request, pk):
    """
    Completely remove a product variant (CartItem) from the user's cart.
    """
    try:
        item = CartItem.objects.get(pk=pk, user=request.user)
    except CartItem.DoesNotExist:
        return Response({"detail": "Not found"}, status=404)

    item.delete()
    return Response({"detail": "Item completely removed from cart."})

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def view_wishlist(request):
    items = WishlistItem.objects.filter(user=request.user).select_related("product", "user")
    return Response(WishlistItemSerializer(items, many=True).data)

@api_view(["POST"])
@permission_classes([IsAuthenticated])
def add_to_wishlist(request):
    product_id = request.data.get("product")
   
    product = get_object_or_404(Product, pk=product_id)
 
    obj, created = WishlistItem.objects.get_or_create(user=request.user, product=product)
    if not created:
        return Response({"detail": "Already in wishlist"})
    return Response(WishlistItemSerializer(obj).data, status=201)

@api_view(["DELETE"])
@permission_classes([IsAuthenticated])
def remove_from_wishlist(request, product_id):
    try:
        item = WishlistItem.objects.get(product_id=product_id, user=request.user)
    except WishlistItem.DoesNotExist:
        return Response({"detail": "Not found"}, status=404)
    item.delete()
    return Response({"detail": "Removed from wishlist"})

# ==========================================================
# ORDERS & INVOICES
# ==========================================================
@api_view(["POST"])
@permission_classes([IsAuthenticated])
def place_order(request):
    serializer = OrderSerializer(data=request.data, context={"request": request})
    if not serializer.is_valid():
        return Response(serializer.errors, status=400)

    order = serializer.save()

    # Calculate price
    subtotal = sum(i.price * i.quantity for i in order.items.all())
    order.total_price = subtotal
    order.status = "pending"
    order.save()
    print("hii")

    # Clear cart
    CartItem.objects.filter(user=request.user).delete()

    # ----------------------------------------------------
    # 🔔 IN-APP NOTIFICATION — BUYER
    # ----------------------------------------------------
    Notification.objects.create(
        user=request.user,
        message=f"Your order #{order.id} has been placed successfully!"
    )
    print("hii2")

    # ----------------------------------------------------
    # 📧 EMAIL — BUYER
    # ----------------------------------------------------
    send_order_confirmation_email(request.user.email, order)
    print("hii3")

    # ----------------------------------------------------
    # 🔔 + 📧 SELLER NOTIFICATIONS
    # ----------------------------------------------------
    seller_ids = order.items.values_list("product__seller_id", flat=True).distinct()
    sellers = User.objects.filter(id__in=seller_ids)

    for seller in sellers:
        # IN-APP notification
        Notification.objects.create(
            user=seller,
            message=f"You received a new order #{order.id}."
        )
        print("hii4")

        # Email
        print(seller.email,order)
        send_seller_new_order_email(seller.email, order)
        print("done")

    return Response({
        "order": OrderSerializer(order).data,
    }, status=201)



@api_view(["GET"])
@permission_classes([IsAuthenticated])
def view_orders(request):
    orders = (
        Order.objects.filter(buyer=request.user)
        .prefetch_related("items__variant", "items__product")
        .select_related("invoice")   
    )
    return Response(OrderSerializer(orders, many=True).data)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def seller_orders(request):
    """
    Lists all orders that contain products belonging to the seller.
    """
    orders = (
        Order.objects.filter(items__product__seller=request.user)
        .distinct()
        .prefetch_related("items__product", "items__variant")
        .order_by("-created_at")
    )
    return Response(OrderSerializer(orders, many=True).data, status=200)


@api_view(["PATCH"])
@permission_classes([IsAuthenticated])
def update_order_status(request, order_id):
    try:
        order = Order.objects.prefetch_related("items__product__seller").get(id=order_id)
    except Order.DoesNotExist:
        return Response({"error": "Order not found"}, status=404)

    if request.user not in [i.product.seller for i in order.items.all()]:
        return Response({"error": "Not authorized"}, status=403)

    new_status = request.data.get("status")

    
    if new_status == "delivered":
        otp = generate_otp()
        order.delivery_otp = otp
        order.is_delivered_verified = False
        order.status = "delivered"
        order.save()

        # EMAIL: OTP
        send_delivery_otp_email(order.buyer.email, otp)

        # IN-APP: Buyer
        Notification.objects.create(
            user=order.buyer,
            message=f"Your order #{order.id} is marked delivered. OTP sent."
        )

        # IN-APP: Seller
        Notification.objects.create(
            user=request.user,
            message=f"OTP sent to buyer for order #{order.id}."
        )

        return Response({
            "message": f"Order #{order.id} marked as delivered. OTP sent."
        }, status=200)


    # General status updates
    serializer = OrderStatusUpdateSerializer(order, data=request.data, partial=True)
    if serializer.is_valid():
        serializer.save()

        # IN-APP: Buyer
        Notification.objects.create(
            user=order.buyer,
            message=f"Your order #{order.id} status updated to '{order.status}'."
        )

        # EMAIL: Buyer
        send_order_status_update_email(order.buyer.email, order)

        # IN-APP: Seller
        for seller in {i.product.seller for i in order.items.all()}:
            Notification.objects.create(
                user=seller,
                message=f"Order #{order.id} status updated to '{order.status}'."
            )

        return Response({"message": f"Order #{order.id} status updated."}, status=200)

    return Response(serializer.errors, status=400)


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def verify_delivery_otp(request, order_id):
    """
    Seller enters OTP. If correct → order becomes completed.
    """
    entered_otp = request.data.get("otp")

    try:
        order = Order.objects.get(id=order_id)
    except Order.DoesNotExist:
        return Response({"error": "Order not found"}, status=404)

    if request.user not in [i.product.seller for i in order.items.all()]:
        return Response({"error": "Not authorized"}, status=403)

    if order.delivery_otp != entered_otp:
        return Response({"error": "Invalid OTP"}, status=400)

    order.status = "completed"
    order.is_delivered_verified = True
    order.delivery_otp = None
    order.save()

    Notification.objects.create(
        user=order.buyer,
        message=f"Your order #{order.id} has been successfully delivered!"
    )

    return Response({
        "message": "OTP verified. Order completed.",
        "status": "completed"
    }, status=200)

@api_view(["POST"])
@permission_classes([IsAuthenticated])
@parser_classes([MultiPartParser, FormParser])
def upload_invoice(request, order_id):
    user = request.user

    try:
        order = Order.objects.get(id=order_id)
    except Order.DoesNotExist:
        return Response({"detail": "Order not found"}, status=404)

    if user.role != "seller":
        return Response({"detail": "Only sellers can upload invoices."}, status=403)
    
    if not request.FILES.get("pdf"):
        return Response({"detail": "Please attach invoice PDF."}, status=400)

    pdf_file = request.FILES["pdf"]

    invoice, created = Invoice.objects.get_or_create(
        order=order,
        defaults={
            "buyer": order.buyer,
            "seller": user,
            "subtotal": order.total_price,
            "total_amount": order.total_price,
            "tax_amount": 0,
            "address": order.address,
        }
    )

    invoice.pdf_file = pdf_file
    invoice.status = "issued"
    invoice.save()

    Notification.objects.create(
        user=order.buyer,
        message=f"Invoice uploaded for Order #{order.id}"
    )

    return Response(
        {
            "message": "Invoice uploaded successfully",
            "invoice_url": invoice.pdf_file.url,
        }
    )

# ==========================================================
# REVIEWS
# ==========================================================
@api_view(["GET", "POST"])
@permission_classes([IsAuthenticated])
def add_review(request, product_id):
    """
    GET → Check eligibility  
    POST → Submit review for a specific product variant.
    """
    product = get_object_or_404(Product, pk=product_id)
    variant_id = request.data.get("variant_id")

    if request.method == "GET":
        eligible = OrderItem.objects.filter(
            order__buyer=request.user,
            product=product,
            variant_id=variant_id,
            order__status__in=["delivered", "processing"]
        ).exists()
        return Response({"eligible": eligible})


    variant = get_object_or_404(ProductVariant, pk=variant_id, product=product)
    has_purchased = OrderItem.objects.filter(
        order__buyer=request.user,
        product=product,
        variant=variant,
        order__status__in=["delivered", "processing"]
    ).exists()

    if not has_purchased:
        return Response({"detail": "You can only review purchased variants."}, status=403)

    data = request.data.copy()
    data["product"] = product.id
    data["variant"] = variant.id
    serializer = ReviewSerializer(data=data, context={"request": request})
    if serializer.is_valid():
        serializer.save(buyer=request.user)
        Notification.objects.create(
            user=product.seller,
            message=f"New review for {product.name} ({variant.variant_label})"
        )
        return Response(serializer.data, status=201)
    return Response(serializer.errors, status=400)


@api_view(["GET"])
@permission_classes([AllowAny])
def list_reviews(request, product_id):
    """
    Returns all reviews for a product or a specific variant.
    """
    variant_id = request.GET.get("variant_id")
    qs = Review.objects.filter(product_id=product_id)
    if variant_id:
        qs = qs.filter(variant_id=variant_id)
    return Response(ReviewSerializer(qs, many=True).data)


# ==========================================================
# NOTIFICATIONS
# ==========================================================
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_notifications(request):
    """
    Retrieve all notifications for the logged-in user.
    """
    notifications = Notification.objects.filter(user=request.user).order_by("-created_at")
    return Response(NotificationSerializer(notifications, many=True).data)


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def mark_notification_read(request, pk):
    """
    Mark a single notification as read.
    """
    try:
        notif = Notification.objects.get(pk=pk, user=request.user)
        notif.is_read = True
        notif.save()
        return Response({"detail": "Notification marked as read"})
    except Notification.DoesNotExist:
        return Response({"detail": "Not found"}, status=404)


# ==========================================================
# QUOTATIONS 
# ==========================================================

@api_view(["POST"])
@permission_classes([IsAuthenticated])
def request_quotation_preproduct(request, product_id):
    product = get_object_or_404(Product, pk=product_id)
    seller = product.seller

    if request.user.role != "buyer":
        return Response({"detail": "Only buyers can request quotations."}, status=403)

    variant_id = request.data.get("variant_id")
    variant = None
    if variant_id:
        variant = get_object_or_404(ProductVariant, pk=variant_id, product=product)

    quantity = int(request.data.get("quantity", 1))

    req, created = QuotationRequest.objects.get_or_create(
        product=product,
        variant=variant,
        buyer=request.user,
        seller=seller,
        quantity=quantity,
    )

    if created:
        variant_text = f" (Variant: {variant.variant_label})" if variant else ""
        Notification.objects.create(
            user=seller,
            message=f"Buyer {request.user.name} requested a quotation for {product.name}{variant_text} (Qty: {quantity})."
        )

    serializer = QuotationRequestSerializer(req, context={"request": request})
    return Response(serializer.data, status=201 if created else 200)



@api_view(["GET"])
@permission_classes([IsAuthenticated])
def list_my_quotation_requests(request):
    if request.user.role == "buyer":
        qs = (
            QuotationRequest.objects.filter(buyer=request.user)
            .select_related("product", "variant", "seller", "quotation")
        )
    else:
        qs = (
            QuotationRequest.objects.filter(seller=request.user)
            .select_related("product", "variant", "buyer", "quotation")
        )

    serializer = QuotationRequestSerializer(qs, many=True, context={"request": request})
    return Response(serializer.data)


@api_view(["POST"])
@permission_classes([IsAuthenticated])
@parser_classes([MultiPartParser, FormParser])
def upload_quotation_for_request(request, request_id):
    """
    Seller uploads a quotation file in response to a pre-order QuotationRequest.
    """
    qreq = get_object_or_404(QuotationRequest, pk=request_id, seller=request.user)

    if qreq.is_resolved:
        return Response({"detail": "Quotation already provided for this request."}, status=400)

    data = request.data.copy()
    data["request"] = qreq.id
    serializer = QuotationSerializer(data=data, context={"request": request})

    if serializer.is_valid():
        quotation = serializer.save(uploaded_by=request.user)
        qreq.is_resolved = True
        qreq.save()

        Notification.objects.create(
            user=qreq.buyer,
            message=f"Seller {request.user.name} uploaded a quotation for {qreq.product.name}."
        )

        return Response(QuotationSerializer(quotation, context={"request": request}).data, status=201)
    return Response(serializer.errors, status=400)



# ==========================================================
# PRODUCT CONVERSATIONS & MESSAGES
# ==========================================================
@api_view(["POST"])
@permission_classes([IsAuthenticated])
@parser_classes([MultiPartParser, FormParser])
def start_or_get_conversation(request):
    """
    Start or fetch a conversation between buyer and seller for an ordered product.
    """
    order_id = request.data.get("order_id")
    product_id = request.data.get("product_id")

    if not order_id or not product_id:
        return Response({"detail": "order_id and product_id are required"}, status=400)

    order = get_object_or_404(Order, pk=order_id, buyer=request.user)
    product = get_object_or_404(Product, pk=product_id)

    if not order.items.filter(product=product).exists():
        return Response({"detail": "Product not part of this order."}, status=400)

    conv, created = ProductConversation.objects.get_or_create(
        order=order,
        product=product,
        buyer=order.buyer,
        seller=product.seller
    )

    if created:
        Notification.objects.create(
            user=product.seller,
            message=f"New conversation started on {product.name} by {order.buyer.name}."
        )

    return Response(ProductConversationSerializer(conv, context={"request": request}).data, status=201 if created else 200)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def list_conversations_for_user(request):
    """
    Fetch all conversations for a user (buyer or seller).
    """
    if request.user.role == "buyer":
        qs = ProductConversation.objects.filter(buyer=request.user).select_related("product", "seller", "order")
    else:
        qs = ProductConversation.objects.filter(seller=request.user).select_related("product", "buyer", "order")

    return Response(ProductConversationSerializer(qs, many=True, context={"request": request}).data)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def list_messages(request, conversation_id):
    """
    Fetch all messages in a specific conversation.
    """
    conv = get_object_or_404(ProductConversation, pk=conversation_id)
    if request.user not in [conv.buyer, conv.seller] and not request.user.is_staff:
        return Response({"detail": "Unauthorized"}, status=403)

    messages = conv.messages.all().order_by("created_at")
    serializer = ProductMessageSerializer(messages, many=True, context={"request": request})

    # Mark unread messages as read
    conv.messages.filter(is_read=False).exclude(sender=request.user).update(is_read=True)

    return Response(serializer.data)


@api_view(["POST"])
@permission_classes([IsAuthenticated])
@parser_classes([MultiPartParser, FormParser])
def send_message(request, conversation_id):
    """
    Send a new message within a product conversation.
    """
    conv = get_object_or_404(ProductConversation, pk=conversation_id)
    if request.user not in [conv.buyer, conv.seller] and not request.user.is_staff:
        return Response({"detail": "Unauthorized"}, status=403)

    data = request.data.copy()
    data["conversation"] = conv.id
    serializer = ProductMessageSerializer(data=data, context={"request": request})

    if serializer.is_valid():
        msg = serializer.save(sender=request.user)
        other_user = conv.seller if request.user == conv.buyer else conv.buyer
        Notification.objects.create(
            user=other_user,
            message=f"New message on {conv.product.name} from {request.user.name}."
        )
        return Response(ProductMessageSerializer(msg, context={"request": request}).data, status=201)

    return Response(serializer.errors, status=400)


# ==========================================================
# SEARCH
# ==========================================================

@api_view(["GET"])
@permission_classes([AllowAny])
def search_products(request):
    """
    Search products by name, CAS No, Ref No, or HSN code.

    Examples:
      /api/search/?q=acetone
      /api/search/?q=67-64-1
      /api/search/?q=P001
      /api/search/?q=2207
    """
    query = request.GET.get("q", "").strip()
    if not query:
        return Response({"detail": "Query parameter 'q' is required."}, status=400)

    # Detect CAS number pattern (e.g., 67-64-1 or 7732-18-5)
    is_cas_no = bool(re.match(r"^\d{2,7}-\d{2}-\d$", query))

    # Build search filter dynamically
    filters = Q()

    if is_cas_no:
        filters |= Q(cas_no__icontains=query)
    else:
        filters |= (
            Q(name__icontains=query)
            | Q(ref_no__icontains=query)
            | Q(hsn__icontains=query)
            | Q(cas_no__icontains=query)
        )

    products = (
        Product.objects.filter(filters)
        .select_related("seller")
        .prefetch_related("variants", "reviews")
        .annotate(
            min_effective_price=Min(
                Coalesce("variants__price", "variants__est_price")
            )
        )
        .order_by("name")
    )

    serializer = ProductSerializer(products, many=True, context={"request": request})
    return Response(serializer.data, status=200)


# ==========================================================
# SELLER DASHBOARD — STATS
# ==========================================================
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def seller_dashboard_stats(request):

    if request.user.role != "seller":
        return Response({"error": "Only sellers can access this endpoint"}, status=403)

    period = request.GET.get("period", "month")
    today = timezone.now().date()

    if period == "day":
        start_date = today
    elif period == "week":
        start_date = today - timedelta(days=7)
    elif period == "month":
        start_date = today - timedelta(days=30)

    else:
        start_date = today.replace(month=1, day=1)

    COMPLETED = Q(order__status="completed") 

    # Basic counts
    total_products = Product.objects.filter(seller=request.user).count()

    total_orders_distinct = (
        Order.objects.filter(
            items__product__seller=request.user,
            status="completed"  
        ).distinct().count()
    )

    # Revenue from completed orders
    revenue_qs = (
        OrderItem.objects.filter(
        Q(product__seller=request.user)
        & Q(order__status="completed")    
        & Q(order__created_at__date__gte=start_date)).annotate(
            eff_price=Coalesce(
                "price",
                Coalesce("variant__price", "variant__est_price"),
                output_field=DecimalField(max_digits=18, decimal_places=2),
            )
        )
    )

    revenue_data = revenue_qs.aggregate(
        total_revenue=Sum(F("quantity") * F("eff_price")),
        total_units=Sum("quantity"),
        total_orders=Count("order", distinct=True),
    )

    # Safe conversions
    total_revenue = float(revenue_data.get("total_revenue") or 0)
    total_units_sold = int(revenue_data.get("total_units") or 0)
    total_orders_placed = int(revenue_data.get("total_orders") or 0)
    avg_order_value = (total_revenue / total_orders_placed) if total_orders_placed else 0

    # Status breakdown (all orders)
    order_status = (
        Order.objects.filter(items__product__seller=request.user)
        .values("status")
        .annotate(count=Count("id"))
        .order_by("status")
    )

    # Recent COMPLETED orders only
    recent_orders = (
        Order.objects.filter(
            items__product__seller=request.user,
            status="completed" 
        )
        .distinct()
        .order_by("-created_at")[:5]
    )

    recent_orders_data = [
        {
            "id": o.id,
            "buyer": o.buyer.name,
            "total_price": float(o.total_price),
            "status": o.status,
            "created_at": o.created_at,
            "items_count": o.items.count(),
        }
        for o in recent_orders
    ]

    return Response(
        {
            "period": period,
            "total_products": total_products,
            "total_orders": total_orders_placed,
            "distinct_orders": total_orders_distinct,
            "total_revenue": total_revenue,
            "total_units_sold": total_units_sold,
            "avg_order_value": round(avg_order_value, 2),
            "order_status_breakdown": list(order_status),
            "recent_orders": recent_orders_data,
        }
    )


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def seller_sales_trends(request):

    if request.user.role != "seller":
        return Response({"error": "Only sellers can access this endpoint"}, status=403)

    period = request.GET.get("period", "month")
    now = timezone.now()
    today = now.date()

    if period == "day":
        start_dt = now - timedelta(hours=24)
        trunc_func = TruncHour("order__created_at")
        time_filter = Q(order__created_at__gte=start_dt)
        fmt = "%Y-%m-%d %H:%M"

    elif period == "week":
        start_date = today - timedelta(days=7)
        trunc_func = TruncDay("order__created_at")
        time_filter = Q(order__created_at__date__gte=start_date)
        fmt = "%Y-%m-%d"

    elif period == "month":
        start_date = today - timedelta(days=30)
        trunc_func = TruncDay("order__created_at")
        time_filter = Q(order__created_at__date__gte=start_date)
        fmt = "%Y-%m-%d"


    else:
        start_date = today.replace(month=1, day=1)
        trunc_func = TruncMonth("order__created_at")
        time_filter = Q(order__created_at__date__gte=start_date)
        fmt = "%Y-%m"

    COMPLETED = Q(order__status="completed")

    base = (
        OrderItem.objects.filter(product__seller=request.user)
        .filter(time_filter & COMPLETED)
        .annotate(
            eff_price=Coalesce(
                "price",
                Coalesce("variant__price", "variant__est_price"),
                output_field=DecimalField(max_digits=18, decimal_places=2),
            ),
            period=trunc_func,
        )
        .values("period")
        .annotate(
            revenue=Sum(F("quantity") * F("eff_price")),
            units_sold=Sum("quantity"),
            order_count=Count("order", distinct=True),
        )
        .order_by("period")
    )

    data = [
        {
            "period": row["period"].strftime(fmt),
            "revenue": float(row["revenue"] or 0),
            "units_sold": int(row["units_sold"] or 0),
            "order_count": int(row["order_count"] or 0),
        }
        for row in base
    ]
   

    return Response({"period": period, "sales_trends": data})

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def seller_product_performance(request):

    if request.user.role != "seller":
        return Response({"error": "Only sellers can access this endpoint"}, status=403)

    period = request.GET.get("period", "month")
    today = timezone.now().date()

    if period == "day":
        start_date = today
    elif period == "week":
        start_date = today - timedelta(days=7)
    elif period == "month":
        start_date = today - timedelta(days=30)
    else:
        start_date = today.replace(month=1, day=1)

    COMPLETED = Q(order__status="completed")

    products_qs = (
        Product.objects.filter(seller=request.user)
        .annotate(
            min_effective_price=Min(
                Coalesce("variants__price", "variants__est_price")
            )
        )
        .prefetch_related("variants", "reviews")
    )

    perf_rows = (
        OrderItem.objects.filter(
            Q(product__seller=request.user)
            & Q(order__status="completed")    
            & Q(order__created_at__date__gte=start_date)
        ).annotate(
            eff_price=Coalesce(
                "price",
                Coalesce("variant__price", "variant__est_price"),
                output_field=DecimalField(max_digits=18, decimal_places=2),
            )
        )
        .values("product_id")
        .annotate(
            total_sales=Sum("quantity"),
            total_revenue=Sum(F("quantity") * F("eff_price")),
            order_count=Count("order", distinct=True),
        )
    )

    by_pid = {
        r["product_id"]: {
            "total_sales": int(r["total_sales"] or 0),
            "total_revenue": float(r["total_revenue"] or 0),
            "order_count": int(r["order_count"] or 0),
        }
        for r in perf_rows
    }

    out_products = []

    for p in products_qs:
        avg_rating = p.reviews.aggregate(avg=Avg("rating")).get("avg") or 0
        review_count = p.reviews.count()

        agg = by_pid.get(p.id, {"total_sales": 0, "total_revenue": 0.0, "order_count": 0})

        out_products.append(
            {
                "id": p.id,
                "name": p.name,
                "category": p.category,
                "min_effective_price": float(p.min_effective_price or 0),
                "total_sales": agg["total_sales"],
                "total_revenue": agg["total_revenue"],
                "order_count": agg["order_count"],
                "average_rating": round(float(avg_rating), 1),
                "review_count": review_count,
            }
        )

   
    category_rows = (
       OrderItem.objects.filter(
            Q(product__seller=request.user)
            & Q(order__status="completed")  
            & Q(order__created_at__date__gte=start_date)
        )

        .annotate(
            eff_price=Coalesce(
                "price",
                Coalesce("variant__price", "variant__est_price"),
                output_field=DecimalField(max_digits=18, decimal_places=2),
            )
        )
        .values("product__category")
        .annotate(
            total_revenue=Sum(F("eff_price") * F("quantity")),
            total_sales=Sum("quantity"),
            product_count=Count("product", distinct=True),
        )
        .order_by("product__category")
    )

    category_performance = [
        {
            "category": row["product__category"],
            "total_revenue": float(row["total_revenue"] or 0),
            "product_count": row["product_count"],
            "total_sales": row["total_sales"],
        }
        for row in category_rows
    ]

    top_sorted = sorted(out_products, key=lambda x: x["total_revenue"], reverse=True)[:10]

    return Response(
        {
            "top_products": top_sorted,
            "category_performance": category_performance,
        }
    )


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def seller_customer_insights(request):

    if request.user.role != "seller":
        return Response({"error": "Only sellers can access this endpoint"}, status=403)

    period = request.GET.get("period", "month")
    today = timezone.now().date()

    if period == "day":
        start_date = today
    elif period == "week":
        start_date = today - timedelta(days=7)
    elif period == "month":
        start_date = today - timedelta(days=30)

    else:
        start_date = today.replace(month=1, day=1)


    customer_data = (
        Order.objects.filter(
            items__product__seller=request.user,
            status="completed",  
            created_at__date__gte=start_date,
        )
        .values("buyer")
        .annotate(
            order_count=Count("id", distinct=True),
            total_spent=Sum("total_price"),
        )
        .order_by("-total_spent")
    )

    total_customers = len(customer_data)
    repeat_customers = sum(1 for c in customer_data if c["order_count"] > 1)
    new_customers = total_customers - repeat_customers

    total_spent = sum(float(c["total_spent"] or 0) for c in customer_data)

    return Response(
        {
            "total_customers": total_customers,
            "repeat_customers": repeat_customers,
            "new_customers": new_customers,
            "repeat_rate": (repeat_customers / total_customers * 100) if total_customers else 0,
            "avg_customer_value": (total_spent / total_customers) if total_customers else 0,
        }
    )